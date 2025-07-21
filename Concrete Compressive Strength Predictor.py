import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

from sklearn.model_selection import train_test_split, RandomizedSearchCV
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_squared_error, r2_score
import joblib

from xgboost import XGBRegressor


# Load the Excel file (make sure it's .xlsx and not open in Excel)
df = pd.read_excel(r"C:\Users\PMYLS\Downloads\Concrete_Data.xlsx", engine='openpyxl')

# Step 1: Print actual column names
print("Original Columns:")
print(df.columns.tolist())

# Step 2: Optional – Rename for easier use
df.columns = [
    "Cement", "Slag", "FlyAsh", "Water", "Superplasticizer",
    "CoarseAggregate", "FineAggregate", "Age", "Strength"
]

# Step 3: Feature Engineering using new column names
df["Water/Cement"] = df["Water"] / (df["Cement"] + 1)
df["Binder"] = df["Cement"] + df["Slag"] + df["FlyAsh"]
df["Fine/Coarse"] = df["FineAggregate"] / (df["CoarseAggregate"] + 1)

# Step 4: Check result
print("\nData Preview with New Features:")
print(df.head())
# Inputs: all except strength
X = df.drop("Strength", axis=1)

# Output: strength
y = df["Strength"]
from sklearn.preprocessing import StandardScaler

scaler = StandardScaler()
X_scaled = scaler.fit_transform(X)
from sklearn.model_selection import train_test_split

X_train, X_test, y_train, y_test = train_test_split(
    X_scaled, y, test_size=0.2, random_state=42
)
from xgboost import XGBRegressor

model = XGBRegressor(
    n_estimators=200,
    learning_rate=0.05,
    max_depth=6,
    subsample=0.9,
    random_state=42
)

model.fit(X_train, y_train)
from sklearn.metrics import mean_squared_error, r2_score
import numpy as np
import matplotlib.pyplot as plt

y_pred = model.predict(X_test)

r2 = r2_score(y_test, y_pred)
rmse = np.sqrt(mean_squared_error(y_test, y_pred))

print(f"✅ R² Score: {r2:.4f}")
print(f"✅ RMSE: {rmse:.4f} MPa")

# Plot predicted vs actual
plt.figure(figsize=(6, 6))
plt.scatter(y_test, y_pred, alpha=0.7)
plt.plot([y.min(), y.max()], [y.min(), y.max()], 'r--')
plt.xlabel("Actual Strength (MPa)")
plt.ylabel("Predicted Strength (MPa)")
plt.title("Actual vs Predicted Strength")
plt.grid(True)
plt.show()
from sklearn.model_selection import RandomizedSearchCV

param_grid = {
    'n_estimators': [100, 200, 300],
    'max_depth': [4, 6, 8, 10],
    'learning_rate': [0.01, 0.05, 0.1],
    'subsample': [0.6, 0.8, 1.0],
    'colsample_bytree': [0.6, 0.8, 1.0]
}

search = RandomizedSearchCV(
    estimator=XGBRegressor(random_state=42),
    param_distributions=param_grid,
    n_iter=30,
    cv=5,
    scoring='neg_mean_squared_error',
    verbose=1,
    n_jobs=-1
)

search.fit(X_train, y_train)

best_model = search.best_estimator_

print("✅ Best Parameters Found:")
print(search.best_params_)
y_pred = best_model.predict(X_test)
r2 = r2_score(y_test, y_pred)
rmse = np.sqrt(mean_squared_error(y_test, y_pred))

print(f"🌟 Tuned R² Score: {r2:.4f}")
print(f"🌟 Tuned RMSE: {rmse:.4f} MPa")
import streamlit as st
import pandas as pd
import numpy as np
import joblib
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.preprocessing import StandardScaler
from xgboost import XGBRegressor
from fpdf import FPDF
import io

# --- Cache model and scaler loading ---
@st.cache_resource
def load_model_and_scaler():
    model = joblib.load("concrete_strength_model.pkl")
    scaler = joblib.load("concrete_scaler.pkl")
    return model, scaler

model, scaler = load_model_and_scaler()

# Load the saved model and scaler
model = joblib.load("concrete_strength_model.pkl")
scaler = joblib.load("concrete_scaler.pkl")


import joblib

joblib.dump(best_model, "concrete_strength_model.pkl")
joblib.dump(scaler, "concrete_scaler.pkl")
print("✅ Model and Scaler saved successfully.")
import streamlit as st
import pandas as pd
import numpy as np
import joblib
import matplotlib.pyplot as plt

# Load model and scaler
model = joblib.load("concrete_strength_model.pkl")
scaler = joblib.load("concrete_scaler.pkl")

st.set_page_config(page_title="Concrete Compressive Strength Predictor", layout="centered")
st.title("🏗️ Concrete Compressive Strength Predictor")
st.markdown("""
This app predicts the **compressive strength (in MPa)** of concrete mixes based on material proportions.

🔹 All input values should be provided in **kg/m³**.
""")

# --- Option Selection ---
option = st.radio("Choose input method:", ["Manual Entry", "Upload CSV File"])

# --- Manual Input ---
if option == "Manual Entry":
    st.subheader("🔢 Enter Mix Proportions")

    cement = st.number_input("Cement (kg/m³)", min_value=0.0)
    slag = st.number_input("Blast Furnace Slag (kg/m³)", min_value=0.0)
    flyash = st.number_input("Fly Ash (kg/m³)", min_value=0.0)
    water = st.number_input("Water (kg/m³)", min_value=0.0)
    superplasticizer = st.number_input("Superplasticizer (kg/m³)", min_value=0.0)
    coarse = st.number_input("Coarse Aggregate (kg/m³)", min_value=0.0)
    fine = st.number_input("Fine Aggregate (kg/m³)", min_value=0.0)
    age = st.number_input("Age (days)", min_value=1)

    if st.button("📊 Predict Strength"):
        water_cement = water / (cement + 1)
        binder = cement + slag + flyash
        fine_coarse = fine / (coarse + 1)

        input_data = [[cement, slag, flyash, water, superplasticizer,
                       coarse, fine, age, water_cement, binder, fine_coarse]]

        scaled_input = scaler.transform(input_data)
        prediction = model.predict(scaled_input)[0]

        st.success(f"✅ Predicted Compressive Strength: {prediction:.2f} MPa")

        ci_lower = prediction - 3.5
        ci_upper = prediction + 3.5
        st.write(f"🔍 **Confidence Interval (±3.5 MPa): {ci_lower:.2f} - {ci_upper:.2f} MPa**")

        # PDF Export
        if st.button("📄 Download Report as PDF"):
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt="Concrete Strength Prediction Report", ln=True, align="C")
            pdf.ln(10)
            fields = ["Cement", "Slag", "Fly Ash", "Water", "Superplasticizer", "Coarse", "Fine", "Age"]
            values = [cement, slag, flyash, water, superplasticizer, coarse, fine, age]
            for f, v in zip(fields, values):
                pdf.cell(200, 10, txt=f"{f}: {v} kg/m³", ln=True)
            pdf.cell(200, 10, txt=f"Predicted Strength: {prediction:.2f} MPa", ln=True)
            pdf.cell(200, 10, txt=f"Confidence Interval: {ci_lower:.2f} - {ci_upper:.2f} MPa", ln=True)
            pdf_output = io.BytesIO()
            pdf.output(pdf_output)
            st.download_button("📥 Download PDF", data=pdf_output.getvalue(), file_name="prediction_report.pdf")

# --- CSV Upload ---
elif option == "Upload CSV File":
    st.subheader("📁 Upload Mix File (CSV)")
    uploaded_file = st.file_uploader("Upload your CSV file", type=["csv"])

    if uploaded_file:
        df = pd.read_csv(uploaded_file)

        try:
            df.columns = [col.strip().title().replace(" ", "") for col in df.columns]
            df.rename(columns={
                "Coarseaggregate": "CoarseAggregate",
                "Fineaggregate": "FineAggregate",
                "Superplasticizer": "Superplasticizer",
                "Flyash": "FlyAsh"
            }, inplace=True)

            df["Water/Cement"] = df["Water"] / (df["Cement"] + 1)
            df["Binder"] = df["Cement"] + df["Slag"] + df["FlyAsh"]
            df["Fine/Coarse"] = df["FineAggregate"] / (df["CoarseAggregate"] + 1)

            X = df[["Cement", "Slag", "FlyAsh", "Water", "Superplasticizer",
                    "CoarseAggregate", "FineAggregate", "Age",
                    "Water/Cement", "Binder", "Fine/Coarse"]]

            scaled_X = scaler.transform(X)
            predictions = model.predict(scaled_X)

            df["Predicted Strength (MPa)"] = predictions
            df["Lower Bound (MPa)"] = predictions - 3.5
            df["Upper Bound (MPa)"] = predictions + 3.5

            st.write("### 📊 Prediction Results:", df)
            csv = df.to_csv(index=False).encode('utf-8')
            st.download_button("📥 Download Results as CSV", data=csv, file_name='predicted_strengths.csv', mime='text/csv')

            # Histogram
            st.write("### 📈 Strength Prediction Distribution")
            fig, ax = plt.subplots()
            ax.hist(predictions, bins=20, color='skyblue', edgecolor='black')
            ax.set_title("Predicted Compressive Strength Distribution")
            ax.set_xlabel("Strength (MPa)")
            ax.set_ylabel("Frequency")
            st.pyplot(fig)

            # Real-time Plot
            st.line_chart(df[["Predicted Strength (MPa)"]])

        except Exception as e:
            st.error(f"❌ Error processing file: {e}")
            # === 🌐 Advanced Features Add-on ===

# 1️⃣ Feature Importance Plot
if st.checkbox("📊 Show Feature Importance"):
    try:
        importances = model.feature_importances_
        feature_names = ["Cement", "Slag", "FlyAsh", "Water", "Superplasticizer",
                         "CoarseAggregate", "FineAggregate", "Age",
                         "Water/Cement", "Binder", "Fine/Coarse"]
        feat_df = pd.DataFrame({'Feature': feature_names, 'Importance': importances})
        feat_df = feat_df.sort_values(by="Importance", ascending=True)

        st.write("### 🧠 Feature Importance (Model Sensitivity)")
        fig, ax = plt.subplots()
        sns.barplot(data=feat_df, x="Importance", y="Feature", ax=ax, palette="viridis")
        st.pyplot(fig)
    except Exception as e:
        st.warning(f"Feature importance plot not available: {e}")

# 2️⃣ Model Accuracy Metrics
st.markdown("### ✅ Model Performance Summary")
st.markdown("- R² Score (Tuned Model): **0.933**")
st.markdown("- RMSE (Root Mean Squared Error): **4.16 MPa**")
st.markdown("- Confidence Interval Range: ±3.5 MPa")

# 3️⃣ Export to Excel
if "df" in locals() or "df" in globals():
    import io
    from openpyxl import Workbook

    excel_output = io.BytesIO()
    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Predictions')

    st.download_button(
        label="📥 Download Predictions as Excel (.xlsx)",
        data=excel_output.getvalue(),
        file_name="concrete_strength_predictions.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# --- Suggestions Box ---
st.markdown("---")
st.info("💡 Tip: Keep Water/Cement ratio < 0.6 for higher strength, and increase binder content for denser concrete.")

# --- Developer Info Section ---
st.markdown("---")
st.markdown("### 👨‍💻 Developer Info")
st.write("""
**Developer:** Arslan Khan  
**Degree:** Final Year Student, B.Sc. Civil Engineering  
**University:** University of Engineering & Technology Lahore, Pakistan  
**Passion:** Dedicated to sustainable civil engineering solutions and integrating them with advanced AI.
""")

# --- Footer ---
st.caption("Made with 💻 by Civil Engineer | Powered by XGBoost + Streamlit")








